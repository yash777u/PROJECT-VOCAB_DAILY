"""
Data generator for DeutschFlash — populates Excel files with all vocabulary data.
Run this script once to create the initial Excel files.
"""
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = [
    "german_word", "pronunciation", "meaning", "example_sentence",
    "option_1", "option_2", "option_3", "option_4",
    "gender", "emoji", "keyword", "note"
]

HEADER_FILL = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="e2e8f0", size=11)
CELL_FONT = Font(name="Calibri", color="cbd5e1", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="334155"),
    right=Side(style="thin", color="334155"),
    top=Side(style="thin", color="334155"),
    bottom=Side(style="thin", color="334155"),
)

# fmt: off
WORD_DATA = {
    1: [
        ("der Vater", "vAH-tah", "the father", "Mein Vater arbeitet viel.", "masculine", "👨", "father dad"),
        ("das Eis", "ice", "the ice cream", "Das Eis schmeckt gut.", "neuter", "🍨", "ice cream sundae"),
        ("die Tür", "t-yur", "the door", "Die Tür ist offen.", "feminine", "🚪", "door wooden"),
        ("das Tor", "tohr", "the gate / barrier", "Das Tor ist geschlossen.", "neuter", "🚧", "gate barrier"),
        ("die Hose", "HOH-zeh", "the trousers / pants", "Die Hose ist blau.", "feminine", "👖", "pants trousers"),
        ("der Stern", "shtern", "the star", "Der Stern leuchtet hell.", "masculine", "⭐", "star space"),
        ("die Sonne", "ZOH-neh", "the sun", "Die Sonne scheint heute.", "feminine", "☀️", "sun sunlight"),
        ("der Mond", "mohnt", "the moon", "Der Mond ist heute voll.", "masculine", "🌙", "moon crescent night"),
        ("der Himmel", "HIM-el", "the sky / heaven", "Der Himmel ist blau.", "masculine", "☁️", "sky clouds"),
        ("der Bleistift", "BLY-shtift", "the pencil", "Ich brauche einen Bleistift.", "masculine", "✏️", "wooden pencil"),
        ("die Pflanze", "PFLAN-tseh", "the plant", "Die Pflanze braucht Wasser.", "feminine", "🪴", "houseplant potted"),
        ("das T-Shirt", "TEE-shurt", "the T-Shirt", "Das T-Shirt ist rot.", "neuter", "👕", "tshirt"),
        ("das Sofa", "ZOH-fah", "the sofa / couch", "Das Sofa ist sehr bequem.", "neuter", "🛋️", "sofa couch"),
        ("der Stift", "shtift", "the pen", "Hast du einen Stift?", "masculine", "🖋️", "fountain pen"),
        ("das Heft", "heft", "the notebook", "Das Heft ist auf dem Tisch.", "neuter", "📒", "notebook spiral"),
        ("der Bonbon", "BONG-bong", "the candy / bonbon", "Der Bonbon ist süß.", "masculine", "🍬", "hard candy"),
        ("der Fisch", "fish", "the fish", "Der Fisch schwimmt im Wasser.", "masculine", "🐟", "fish swimming"),
        ("die Blume", "BLOO-meh", "the flower", "Die Blume ist schön.", "feminine", "🌸", "flower blossom"),
        ("der Ball", "bahl", "the ball", "Der Ball ist rund.", "masculine", "⚽", "soccer ball"),
        ("die Lampe", "LAHM-peh", "the lamp", "Die Lampe ist an.", "feminine", "💡", "lamp lightbulb"),
    ],
    2: [
        ("Tschüss", "tshuess", "Bye", "Tschüss, bis morgen!", "phrase", "✌️", "wave goodbye"),
        ("Auf Wiedersehen", "owf VEE-der-zay-en", "Goodbye", "Auf Wiedersehen, Herr Müller.", "phrase", "🤝", "handshake professional"),
        ("Gute Nacht", "GOO-teh nakht", "Good night", "Gute Nacht, schlaf gut!", "phrase", "🌙", "sleeping bed night"),
        ("Bis morgen", "bis MOR-gen", "See you tomorrow", "Bis morgen in der Schule!", "phrase", "📅", "calendar schedule"),
        ("Bis abend", "bis AH-bent", "See you this evening", "Bis abend, mein Freund!", "phrase", "🌇", "sunset dusk"),
        ("Danke / Danke schön", "DAN-keh / DAN-keh shoen", "Thank you / Thank you very much", "Danke schön für die Hilfe!", "phrase", "🙏", "thank you sign"),
        ("Bitte schön", "BIT-teh shoen", "You're welcome", "Bitte schön, gern geschehen!", "phrase", "😊", "happy smile face"),
        ("Schönen Tag noch", "SHOEN-en tahk nokh", "Have a nice day", "Schönen Tag noch, Frau Schmidt!", "phrase", "🌤️", "pleasant day park"),
        ("Schönen Abend noch", "SHOEN-en AH-bent nokh", "Have a nice evening", "Schönen Abend noch!", "phrase", "🌆", "city lights night"),
        ("Gleichfalls", "GLYKH-fals", "Likewise / Same to you", "Danke, gleichfalls!", "phrase", "🤝", "agreement matching"),
        ("Wie geht es dir?", "vee gayt es deer", "How are you? (Informal)", "Hallo! Wie geht es dir?", "phrase", "👫", "friends talking"),
        ("Wie geht's?", "vee gayts", "How's it going?", "Hey, wie geht's?", "phrase", "⚡", "casual talk greeting"),
        ("Wie geht es Ihnen?", "vee gayt es EE-nen", "How are you? (Formal)", "Guten Tag, wie geht es Ihnen?", "phrase", "👔", "business interview meeting"),
        ("Gut", "goot", "Good", "Mir geht es gut.", "phrase", "👍", "thumbs up positive"),
        ("Sehr gut", "zair goot", "Very good", "Mir geht es sehr gut!", "phrase", "🌟", "happy celebrating star"),
        ("Prima / Super / Wunderbar", "PREE-mah / ZOO-per / VOON-der-bahr", "Great / Wonderful", "Das ist wunderbar!", "phrase", "🤩", "excited jumping joy"),
        ("Ausgezeichnet", "ows-gay-tsy-khnet", "Excellent", "Die Arbeit ist ausgezeichnet.", "phrase", "💎", "gold medal trophy"),
        ("So lala", "zo lah-lah", "So-so", "Wie geht's? — So lala.", "phrase", "😐", "indifferent expression"),
        ("Nicht so gut", "nikht zo goot", "Not so good", "Mir geht es nicht so gut.", "phrase", "😕", "unhappy sad face"),
        ("Schlecht", "shlekht", "Bad", "Das Wetter ist schlecht.", "phrase", "👎", "thumbs down negative"),
        ("Es geht mir...", "es gayt meer", "I am doing...", "Es geht mir gut, danke!", "phrase", "📝", "writing message paper"),
        ("Hallo", "hah-loh", "Hello", "Hallo, ich bin Anna!", "phrase", "🖐️", "greeting wave"),
        ("Guten Morgen", "GOO-ten MOR-gen", "Good morning", "Guten Morgen, wie geht's?", "phrase", "🌅", "sunrise coffee"),
        ("Guten Tag", "GOO-ten TAHK", "Good day / Hello", "Guten Tag, Herr Meier.", "phrase", "☀️", "bright sunny landscape"),
        ("Guten Abend", "GOO-ten AH-bent", "Good evening", "Guten Abend, willkommen!", "phrase", "🌆", "sunset skyline"),
    ],
    3: [
        ("das Auto", "OW-toh", "the car", "Das Auto ist schnell.", "neuter", "🚗", "modern automobile sportscar"),
        ("die Schokolade", "shoh-koh-LAH-deh", "the chocolate", "Die Schokolade ist lecker.", "feminine", "🍫", "chocolate bar candy"),
        ("der Baum", "bowm", "the tree", "Der Baum ist sehr groß.", "masculine", "🌳", "green forest tree"),
        ("der Hund", "hoont", "the dog", "Der Hund spielt im Garten.", "masculine", "🐕", "dog puppy labrador"),
        ("die Katze", "KAHT-seh", "the cat", "Die Katze schläft gern.", "feminine", "🐈", "cat domestic kitten"),
        ("die Uhr", "oor", "the clock / watch", "Die Uhr zeigt drei Uhr.", "feminine", "⏰", "alarm clock watch"),
        ("das Zimmer", "TSIM-er", "the room", "Das Zimmer ist sauber.", "neuter", "🏠", "bedroom livingroom"),
        ("der Lehrer", "LAY-rer", "the teacher (male)", "Der Lehrer erklärt die Aufgabe.", "masculine", "👨‍🏫", "male teacher blackboard"),
        ("die Lehrerin", "LAY-reh-rin", "the teacher (female)", "Die Lehrerin ist sehr nett.", "feminine", "👩‍🏫", "female teacher classroom"),
        ("der Tisch", "tish", "the table", "Der Tisch ist aus Holz.", "masculine", "🪵", "wooden desk dining table"),
        ("der Stuhl", "shtool", "the chair", "Der Stuhl ist kaputt.", "masculine", "🪑", "wood design chair"),
        ("der Kuli", "KOO-lee", "the ballpoint pen", "Ich schreibe mit dem Kuli.", "masculine", "🖊️", "ballpoint writing pen"),
        ("das Fenster", "FEN-ster", "the window", "Das Fenster ist offen.", "neuter", "🪟", "window frame daylight"),
        ("der Computer", "kom-PEW-ter", "the computer", "Der Computer ist neu.", "masculine", "💻", "desktop computer keyboard"),
        ("der Mann", "mahn", "the man", "Der Mann liest ein Buch.", "masculine", "👨", "man smiling portrait"),
        ("die Frau", "frow", "the woman", "Die Frau trinkt Kaffee.", "feminine", "👩", "woman smiling model"),
        ("das Kind", "kint", "the child", "Das Kind spielt draußen.", "neuter", "🧒", "happy child play"),
        ("das Haus", "hows", "the house", "Das Haus hat einen Garten.", "neuter", "🏠", "suburban brick house"),
        ("der Zug", "tsook", "the train", "Der Zug kommt um zehn Uhr.", "masculine", "🚂", "modern passenger train"),
    ],
    4: [
        ("der Norden", "Nor-den", "North", "Hamburg liegt im Norden.", "masculine", "⬆️", "compass north directional"),
        ("der Osten", "Os-ten", "East", "Berlin liegt im Osten.", "masculine", "➡️", "compass direction east sunrise"),
        ("der Süden", "Zue-den", "South", "München liegt im Süden.", "masculine", "⬇️", "tropical island south beach"),
        ("der Westen", "Ves-ten", "West", "Köln liegt im Westen.", "masculine", "⬅️", "west sunset ocean horizon"),
        ("der Frühling", "Froo-ling", "Spring", "Im Frühling blühen die Blumen.", "masculine", "🌸", "cherry blossom field park"),
        ("der Sommer", "Zom-mer", "Summer", "Im Sommer ist es heiß.", "masculine", "☀️", "sunny beach pool holiday"),
        ("der Herbst", "Hairpst", "Autumn / Fall", "Im Herbst fallen die Blätter.", "masculine", "🍂", "autumn dry forest leaves"),
        ("der Winter", "Vin-ter", "Winter", "Im Winter schneit es oft.", "masculine", "❄️", "snow mountain cold cottage"),
        ("der Januar", "Yah-nu-ahr", "January", "Im Januar ist es kalt.", "masculine", "❄️", "calendar new year frost"),
        ("der Februar", "Fay-bru-ahr", "February", "Februar hat 28 Tage.", "masculine", "❄️", "valentine february heart"),
        ("der März", "Mairts", "March", "Im März wird es wärmer.", "masculine", "🌸", "spring march daffodil"),
        ("der April", "Ah-pril", "April", "Im April regnet es oft.", "masculine", "🌸", "rain shower umbrella april"),
        ("der Mai", "My", "May", "Im Mai sind die Felder grün.", "masculine", "🌸", "may green fields landscape"),
        ("der Juni", "Yu-nee", "June", "Im Juni beginnt der Sommer.", "masculine", "☀️", "june summer sunflowers"),
        ("der Juli", "Yu-lee", "July", "Im Juli ist es sehr warm.", "masculine", "☀️", "fireworks july night"),
        ("der August", "Ow-goost", "August", "Im August fahren wir ans Meer.", "masculine", "☀️", "swimming pool float august"),
        ("der September", "Zep-tem-ber", "September", "Im September beginnt die Schule.", "masculine", "🍂", "school classroom september notebook"),
        ("der Oktober", "Ok-to-ber", "October", "Im Oktober ist Oktoberfest.", "masculine", "🍂", "halloween orange pumpkins october"),
        ("der November", "No-vem-ber", "November", "Im November ist es neblig.", "masculine", "🍂", "thanksgiving cozy turkey"),
        ("der Dezember", "De-tsem-ber", "December", "Im Dezember feiern wir Weihnachten.", "masculine", "❄️", "christmas fireplace gifts tree"),
        ("Ich", "ikh", "I", "Ich bin Student.", "pronoun", "👤", "myself individual portrait"),
        ("Du", "doo", "you (informal)", "Du bist mein Freund.", "pronoun", "🫵", "pointing finger look at you"),
        ("Er", "air", "he", "Er spielt Fußball.", "pronoun", "👨", "man business suit model"),
        ("Sie", "zee", "she", "Sie liest ein Buch.", "pronoun", "👩", "woman portrait smiling lifestyle"),
        ("Es", "es", "it", "Es regnet heute.", "pronoun", "🤖", "robot technology machine"),
        ("Wir", "veer", "we", "Wir gehen ins Kino.", "pronoun", "👥", "group friends happy hugging"),
        ("Ihr", "eer", "you all (informal plural)", "Ihr seid willkommen.", "pronoun", "👥", "classroom students listening"),
        ("Sie (formal)", "zee", "they / You (formal)", "Sie kommen morgen.", "pronoun", "🤝", "formal business handshake interview"),
    ],
    5: [
        ("kochen", "kokh-en", "to cook", "Ich koche Suppe.", "verb", "🍲", "cooking chef stove pot kitchen"),
        ("machen", "makh-en", "to do/make", "Ich mache Hausaufgaben.", "verb", "✍️", "doing homework writing"),
        ("suchen", "zookh-en", "to search/look for", "Ich suche mein Buch.", "verb", "📖", "searching magnifying glass map"),
        ("trinken", "trink-en", "to drink", "Ich trinke Wasser.", "verb", "💧", "drinking glass water sip"),
        ("hören", "hoer-en", "to hear/listen", "Ich höre Musik.", "verb", "🎧", "headphones listening music girl"),
        ("fragen", "frah-gen", "to ask", "Ich frage den Lehrer.", "verb", "🙋‍♂️", "raising hand classroom pupil"),
        ("lieben", "leeb-en", "to love", "Ich liebe meine Familie.", "verb", "❤️", "heart love couples"),
        ("trinken (Wasser)", "trink-en (vah-ser)", "to drink (water)", "Ich trinke viel Wasser.", "verb", "💧", "drinking clean glass water"),
        ("trinken (Bier)", "trink-en (beer)", "to drink (beer)", "Er trinkt gern Bier.", "verb", "🍺", "german beer stein tankard pint"),
        ("Wie spät ist es?", "vee shpayt ist es", "How late is it?", "Wie spät ist es jetzt?", "phrase", "🕒", "antique wall clock"),
        ("Wie viel Uhr ist es?", "vee feel oor ist es", "What time is it?", "Wie viel Uhr ist es bitte?", "phrase", "⏰", "elegant wrist watch design"),
        ("Uhr", "oor", "Clock/o'clock", "Es ist drei Uhr.", "time", "🕰️", "clock dials hands"),
        ("Minuten", "mi-noo-ten", "Minute", "Es dauert fünf Minuten.", "time", "⏳", "minute sand stopwatch hourglass"),
        ("Vor", "for", "Before (to the hour)", "Es ist zehn vor drei.", "time", "⬅️", "back arrow reverse navigation"),
        ("Nach", "nahkh", "After (past the hour)", "Es ist fünf nach zwei.", "time", "➡️", "forward arrow next icon"),
        ("Viertel nach", "feer-tel nahkh", "15 Minutes Past", "Es ist Viertel nach eins.", "time", "🕒", "clock quarter past three"),
        ("Halb", "hahlp", "30 Minutes Past (Half)", "Es ist halb drei.", "time", "🌓", "moon phase half slice crescent"),
        ("Viertel vor", "feer-tel for", "15 Minutes To", "Es ist Viertel vor vier.", "time", "⏳", "clock quarter to nine hourglass"),
        ("Dreiviertel", "dry-feer-tel", "45 Minutes Past", "Es ist Dreiviertel zwölf.", "time", "🥧", "pie chart three quarters diagram"),
    ],
}
# fmt: on

NOTE_MAP = {
    "masculine": "Noun • Singular",
    "feminine": "Noun • Singular",
    "neuter": "Noun • Singular",
    "phrase": "Phrase • Conversation",
    "pronoun": "Pronoun",
    "verb": "Verb",
    "time": "Time Component",
}

DAY_TITLES = {
    1: "Essential Nouns I",
    2: "Social & Greetings",
    3: "Essential Nouns II",
    4: "Directions, Seasons, Months & Pronouns",
    5: "Verbs, Time Phrases & Telling Time",
}


def _all_meanings():
    """Collect every meaning across all days for distractor generation."""
    meanings = []
    for day_words in WORD_DATA.values():
        for w in day_words:
            meanings.append(w[2])
    return meanings


def _generate_options(correct_meaning: str, all_meanings: list) -> list:
    """Return a shuffled list of 4 options including the correct answer."""
    distractors = [m for m in all_meanings if m != correct_meaning]
    chosen = random.sample(distractors, min(3, len(distractors)))
    options = [correct_meaning] + chosen
    random.shuffle(options)
    return options


def create_level_workbook(filepath: str, word_data: dict):
    """Create a structured Excel workbook for a vocabulary level."""
    wb = Workbook()
    all_meanings = _all_meanings()

    for day_num in sorted(word_data.keys()):
        title = f"Day {day_num}"
        if day_num == 1:
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title=title)

        # Write headers
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # Write word data
        for row_idx, word_tuple in enumerate(word_data[day_num], 2):
            word, pron, meaning, example, gender, emoji, keyword = word_tuple
            note = NOTE_MAP.get(gender, "Phrase • Conversation")
            options = _generate_options(meaning, all_meanings)

            row_data = [
                word, pron, meaning, example,
                options[0], options[1], options[2], options[3],
                gender, emoji, keyword, note,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = CELL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Auto-adjust column widths
        for col_idx in range(1, len(COLUMNS) + 1):
            max_len = len(COLUMNS[col_idx - 1])
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 45)

    wb.save(filepath)
    print(f"  ✅ Created: {filepath}")


def create_empty_template(filepath: str):
    """Create an empty Excel template with proper headers and one sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Day 1"

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

    wb.save(filepath)
    print(f"  📄 Created template: {filepath}")


# ═══════════════════════════════════════════════════════════════
# Entertainment & Subtitles Level — learn German via movies/TV/music
# ═══════════════════════════════════════════════════════════════
# fmt: off
ENTERTAINMENT_DATA = {
    1: [
        # Day 1: Movie & TV Show Phrases (common subtitle lines)
        ("Hilfe!", "HIL-feh", "Help!", "Hilfe! Ruft die Polizei!", "phrase", "🆘", "emergency rescue alarm"),
        ("Was ist los?", "vas ist lohs", "What's going on?", "Was ist los mit dir?", "phrase", "❓", "confused surprise reaction"),
        ("Keine Ahnung", "KY-neh AH-nung", "No idea", "Ich habe keine Ahnung.", "phrase", "🤷", "shrug confused person"),
        ("Genau", "geh-NOW", "Exactly", "Genau, das stimmt!", "phrase", "🎯", "bullseye target right"),
        ("Wahnsinn!", "VAHN-zinn", "Insane! / Crazy!", "Das ist Wahnsinn!", "phrase", "🤯", "mind blown explosion"),
        ("Lass mich in Ruhe", "lass mikh in ROO-eh", "Leave me alone", "Bitte, lass mich in Ruhe!", "phrase", "🚫", "stop sign privacy"),
        ("Es tut mir leid", "es toot meer lyte", "I'm sorry", "Es tut mir leid, das war mein Fehler.", "phrase", "😔", "sorry apology regret"),
        ("Ich verstehe nicht", "ikh fair-SHTAY-eh nikht", "I don't understand", "Tut mir leid, ich verstehe nicht.", "phrase", "😕", "confused question face"),
        ("Mach schnell!", "mahkh shnell", "Hurry up!", "Mach schnell, wir sind spät!", "phrase", "⏩", "running fast speed"),
        ("Das ist verrückt", "das ist fair-RUEKT", "That's crazy", "Das ist total verrückt!", "phrase", "😜", "wild crazy party"),
        ("Sei vorsichtig!", "zye FOR-zikh-tikh", "Be careful!", "Sei vorsichtig auf der Straße!", "phrase", "⚠️", "warning caution sign"),
        ("Ich bin müde", "ikh bin MUE-deh", "I'm tired", "Ich bin so müde heute.", "phrase", "😴", "sleeping tired yawning"),
        ("Das ist langweilig", "das ist LANG-vy-likh", "That's boring", "Der Film ist langweilig.", "phrase", "😑", "bored yawning face"),
        ("Unglaublich!", "UN-gloub-likh", "Unbelievable!", "Das ist unglaublich schön!", "phrase", "😲", "amazed astonished awe"),
        ("Ich habe Hunger", "ikh HAH-beh HUNG-er", "I'm hungry", "Ich habe großen Hunger.", "phrase", "🍕", "hungry food craving"),
        ("Aufpassen!", "OWF-pah-sen", "Watch out!", "Pass auf, ein Auto kommt!", "phrase", "👀", "eyes watching alert"),
        ("Alles klar", "AH-les klahr", "All clear / Got it", "Alles klar, ich komme!", "phrase", "👌", "ok approval check"),
        ("Stimmt das?", "shtimt das", "Is that right?", "Du bist 30? Stimmt das?", "phrase", "🧐", "detective investigating curious"),
        ("Ich bin dran", "ikh bin drahn", "It's my turn", "Warte, ich bin dran!", "phrase", "☝️", "pointing up me next"),
        ("Ach so!", "akh zoh", "Oh I see!", "Ach so! Jetzt verstehe ich.", "phrase", "💡", "lightbulb understanding idea"),
    ],
    2: [
        # Day 2: Emotions & Reactions (heard in dramas/reality TV)
        ("Ich freue mich", "ikh FROY-eh mikh", "I'm happy / I'm looking forward", "Ich freue mich auf das Konzert!", "phrase", "😊", "happy excited joyful"),
        ("Ich habe Angst", "ikh HAH-beh angst", "I'm afraid / scared", "Ich habe Angst vor Spinnen.", "phrase", "😨", "scared afraid horror"),
        ("Das nervt", "das nairft", "That's annoying", "Das nervt mich total!", "phrase", "😤", "angry irritated annoyed"),
        ("Ich vermisse dich", "ikh fair-MIS-seh dikh", "I miss you", "Ich vermisse dich so sehr!", "phrase", "💔", "miss you heartache sad"),
        ("Du spinnst!", "doo shpinst", "You're crazy!", "Du spinnst doch!", "phrase", "🌀", "dizzy spiral confusion"),
        ("Ich schwöre", "ikh SHVOE-reh", "I swear", "Ich schwöre, das ist wahr!", "phrase", "🤞", "promise swear honest"),
        ("Auf keinen Fall!", "owf KY-nen fahl", "No way!", "Auf keinen Fall mache ich das!", "phrase", "🙅", "refuse denial no way"),
        ("Mir ist egal", "meer ist eh-GAHL", "I don't care", "Mir ist das völlig egal.", "phrase", "💅", "whatever indifferent casual"),
        ("Krass!", "krahs", "Awesome! / Intense!", "Das war echt krass!", "phrase", "🔥", "fire awesome intense cool"),
        ("Geil!", "gyle", "Cool! / Awesome!", "Das neue Lied ist geil!", "phrase", "🤙", "cool awesome surfer"),
        ("Ich bin sauer", "ikh bin ZOW-er", "I'm angry/upset", "Ich bin sauer auf dich!", "phrase", "😡", "angry red furious face"),
        ("Das geht nicht", "das gayt nikht", "That won't work", "Nein, das geht so nicht!", "phrase", "❌", "cross error wrong"),
        ("Zum Glück!", "tsum gluek", "Luckily!", "Zum Glück hast du angerufen!", "phrase", "🍀", "four leaf clover luck"),
        ("Ich bin stolz", "ikh bin shtolts", "I'm proud", "Ich bin stolz auf dich!", "phrase", "🏆", "trophy proud achievement"),
        ("Halt die Klappe!", "hahlt dee KLAH-peh", "Shut up!", "Halt die Klappe, bitte!", "phrase", "🤫", "shush quiet silence"),
        ("Wie peinlich!", "vee PINE-likh", "How embarrassing!", "Wie peinlich war das denn!", "phrase", "🫣", "embarrassed hiding face"),
        ("Ich glaube dir", "ikh GLOW-beh deer", "I believe you", "Okay, ich glaube dir.", "phrase", "🤝", "trust handshake agreement"),
        ("Das macht Spaß!", "das makht shpahs", "That's fun!", "Tanzen macht total Spaß!", "phrase", "🎉", "party confetti celebration"),
    ],
    3: [
        # Day 3: Streaming, Social Media & Pop Culture Slang
        ("der Film", "film", "the movie/film", "Der Film war spannend.", "masculine", "🎬", "movie cinema popcorn"),
        ("die Serie", "ZEH-ree-eh", "the series/show", "Die Serie hat zehn Folgen.", "feminine", "📺", "television streaming binge"),
        ("die Folge", "FOL-geh", "the episode", "Die nächste Folge kommt morgen.", "feminine", "▶️", "play button next episode"),
        ("der Untertitel", "UN-ter-tee-tel", "the subtitle", "Ich lese die Untertitel.", "masculine", "💬", "caption subtitle text"),
        ("die Sendung", "ZEN-dung", "the show/broadcast", "Die Sendung beginnt um 20 Uhr.", "feminine", "📡", "broadcast satellite antenna"),
        ("der Schauspieler", "SHOW-shpee-ler", "the actor", "Der Schauspieler spielt gut.", "masculine", "🎭", "actor theatre mask"),
        ("die Schauspielerin", "SHOW-shpee-leh-rin", "the actress", "Die Schauspielerin ist berühmt.", "feminine", "🌟", "famous star celebrity"),
        ("das Lied", "leet", "the song", "Das Lied ist wunderschön.", "neuter", "🎵", "music song melody"),
        ("der Sänger", "ZENG-er", "the singer (male)", "Der Sänger singt laut.", "masculine", "🎤", "male singer microphone"),
        ("die Sängerin", "ZENG-eh-rin", "the singer (female)", "Die Sängerin hat eine tolle Stimme.", "feminine", "🎙️", "female singer performer"),
        ("die Musik", "moo-ZEEK", "the music", "Ich höre gern Musik.", "feminine", "🎶", "music notes headphones"),
        ("das Konzert", "kon-TSAIRT", "the concert", "Das Konzert war fantastisch!", "neuter", "🎸", "concert stage guitar"),
        ("der Kanal", "kah-NAHL", "the channel", "Welcher Kanal zeigt das Spiel?", "masculine", "📲", "youtube channel subscribe"),
        ("streamen", "STREE-men", "to stream", "Wir streamen den Film heute.", "verb", "📡", "streaming laptop screen"),
        ("liken", "LY-ken", "to like (social media)", "Ich like dein Foto!", "verb", "👍", "thumbs up social media"),
        ("teilen", "TY-len", "to share", "Kannst du das Video teilen?", "verb", "🔗", "share link forward"),
        ("abonnieren", "ah-bo-NEE-ren", "to subscribe", "Bitte abonniert meinen Kanal!", "verb", "🔔", "subscribe bell notification"),
        ("die Werbung", "VAIR-bung", "the advertisement", "Die Werbung ist zu lang.", "feminine", "📢", "advertisement commercial billboard"),
        ("spannend", "SHPAH-nend", "exciting/thrilling", "Der Krimi ist sehr spannend!", "phrase", "😬", "suspense thriller popcorn"),
        ("lustig", "LUS-tikh", "funny", "Die Komödie ist sehr lustig!", "phrase", "😂", "laughing funny comedy"),
    ],
}
# fmt: on

ENTERTAINMENT_DAY_TITLES = {
    1: "Movie & TV Phrases",
    2: "Emotions & Reactions",
    3: "Streaming & Pop Culture",
}


def generate_all():
    """Main entry point — creates all Excel files."""
    data_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
    os.makedirs(data_dir, exist_ok=True)

    print("\n🇩🇪 DeutschFlash Data Generator")
    print("=" * 40)

    # A1 — fully populated
    a1_path = os.path.join(data_dir, "A1_vocab.xlsx")
    print("\n📚 Generating A1 vocabulary (111 words, 5 days)...")
    create_level_workbook(a1_path, WORD_DATA)

    # ENT (Entertainment & Subtitles) — fully populated
    ent_path = os.path.join(data_dir, "ENT_vocab.xlsx")
    ent_total = sum(len(v) for v in ENTERTAINMENT_DATA.values())
    print(f"\n🎬 Generating Entertainment & Subtitles vocabulary ({ent_total} words, {len(ENTERTAINMENT_DATA)} days)...")
    create_level_workbook(ent_path, ENTERTAINMENT_DATA)

    # A2, B1, B2 — empty templates
    for level in ["A2", "B1", "B2"]:
        tpl_path = os.path.join(data_dir, f"{level}_vocab.xlsx")
        print(f"\n📝 Generating {level} template...")
        create_empty_template(tpl_path)

    total = sum(len(v) for v in WORD_DATA.values()) + ent_total
    print(f"\n🎉 Done! {total} words generated across A1 + ENT levels.")
    print(f"📂 Files saved to: {data_dir}\n")


if __name__ == "__main__":
    generate_all()

