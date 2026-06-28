from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"


def main() -> int:
    required = [
        ROOT / "app.py",
        ROOT / "requirements.txt",
        DATA / "German.pptx",
        DATA / "A1_vocab.xlsx",
        DATA / "Movie_vocab.xlsx",
        DATA / "TOPICS_WISE_vocab.xlsx",
    ]
    missing = [str(path.relative_to(ROOT)) for path in required if not path.exists()]
    if missing:
        print("Missing required files:")
        for item in missing:
            print(f"- {item}")
        return 1

    workbook_count = 0
    populated_rows = 0
    for workbook in sorted(DATA.glob("*_vocab.xlsx")):
        if workbook.name.startswith((".~", "~$")):
            continue
        wb = load_workbook(workbook, read_only=True, data_only=True)
        workbook_count += 1
        for sheet in wb.worksheets:
            populated_rows += max(sheet.max_row - 1, 0)
        wb.close()

    slide_count = len(list((DATA / "slides_cache").glob("slide-*.png")))
    audio_count = len(list((DATA / "audio_cache").rglob("*.mp3")))

    print("DeutschFlash deploy folder looks ready.")
    print(f"Workbooks: {workbook_count}")
    print(f"Vocabulary rows: {populated_rows}")
    print(f"Cached slides: {slide_count}")
    print(f"Cached audio files: {audio_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
